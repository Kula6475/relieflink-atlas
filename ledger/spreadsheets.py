"""Spreadsheet bridge: import a partner's inventory sheet, export a live-linked workbook.

Import accepts .xlsx or .csv with flexible headers (site/location, category/type,
count/qty, optional county/lat/lon). Unknown sites are created on the fly, rows land
as snapshots with source="spreadsheet".

Export builds a fresh workbook straight from the live ledger every time it is
downloaded, so the file a partner gets back is always in sync: same URL, current data.
"""

import csv
import io
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from ledger.models import InventorySnapshot, Site
from ledger.queries import compute_gaps, latest_forecasts, latest_snapshots
from shared.config import CATEGORIES

COLUMN_ALIASES = {
    "site": "site",
    "site name": "site",
    "location": "site",
    "pantry": "site",
    "name": "site",
    "category": "category",
    "item category": "category",
    "type": "category",
    "count": "count",
    "qty": "count",
    "quantity": "count",
    "units": "count",
    "stock": "count",
    "amount": "count",
    "county": "county",
    "lat": "lat",
    "latitude": "lat",
    "lon": "lon",
    "lng": "lon",
    "longitude": "lon",
}

CATEGORY_ALIASES = {
    "canned goods": "canned_goods",
    "canned": "canned_goods",
    "cans": "canned_goods",
    "canned food": "canned_goods",
    "fresh produce": "produce",
    "fruit": "produce",
    "vegetables": "produce",
    "dry goods": "dry_goods",
    "dry": "dry_goods",
    "grains": "dry_goods",
    "pantry staples": "dry_goods",
}

# Where a spreadsheet-created site lands on the map when no lat/lon column is given.
DEFAULT_LOCATION = (37.5, -121.9)

EXPORT_NOTE = (
    "This workbook is generated live from the ReliefLink shared ledger. "
    "Re-download it any time from /spreadsheets/export to get current data. "
    "To push YOUR counts into the ledger, fill the Upload sheet's format and "
    "upload it at the dashboard's Spreadsheets tab (or POST /spreadsheets/import)."
)


def normalize_category(raw: str) -> str | None:
    value = str(raw).strip().lower().replace("-", " ")
    if value.replace(" ", "_") in CATEGORIES:
        return value.replace(" ", "_")
    return CATEGORY_ALIASES.get(value)


def rows_from_upload(filename: str, data: bytes) -> list[dict]:
    """Parse an uploaded .xlsx or .csv into dicts keyed by canonical column names."""
    if filename.lower().endswith(".csv"):
        reader = csv.reader(io.StringIO(data.decode("utf-8-sig")))
        raw_rows = [row for row in reader if any(str(cell).strip() for cell in row)]
    else:
        sheet = load_workbook(io.BytesIO(data), read_only=True, data_only=True).active
        raw_rows = [
            [cell if cell is not None else "" for cell in row]
            for row in sheet.iter_rows(values_only=True)
            if any(cell not in (None, "") for cell in row)
        ]

    if not raw_rows:
        return []

    header = [COLUMN_ALIASES.get(str(cell).strip().lower(), None) for cell in raw_rows[0]]
    rows = []
    for raw in raw_rows[1:]:
        row = {key: raw[i] for i, key in enumerate(header) if key and i < len(raw)}
        if row:
            rows.append(row)
    return rows


def import_rows(rows: list[dict], session: Session) -> dict:
    """Turn parsed rows into snapshots. Creates unknown sites. Returns a summary."""
    sites_by_name = {site.name.strip().lower(): site for site in session.exec(select(Site)).all()}
    imported, created_sites, skipped = 0, [], []

    for index, row in enumerate(rows, start=2):  # 2 = first data row in the sheet
        site_name = str(row.get("site", "")).strip()
        category = normalize_category(row.get("category", ""))
        try:
            count = int(float(row.get("count", "")))
        except (TypeError, ValueError):
            count = None

        if not site_name or category is None or count is None or count < 0:
            skipped.append({"row": index, "reason": "need site, a known category, and a count"})
            continue

        site = sites_by_name.get(site_name.lower())
        if site is None:
            try:
                lat = float(row.get("lat", "") or DEFAULT_LOCATION[0])
                lon = float(row.get("lon", "") or DEFAULT_LOCATION[1])
            except (TypeError, ValueError):
                lat, lon = DEFAULT_LOCATION
            site = Site(
                name=site_name,
                county=str(row.get("county", "") or "Unknown").strip(),
                lat=lat,
                lon=lon,
            )
            session.add(site)
            session.commit()
            session.refresh(site)
            sites_by_name[site_name.lower()] = site
            created_sites.append(site.name)

        session.add(
            InventorySnapshot(
                site_id=site.id,
                category=category,
                count=count,
                confidence=1.0,
                source="spreadsheet",
            )
        )
        imported += 1

    session.commit()
    return {"imported_rows": imported, "created_sites": created_sites, "skipped": skipped}


def _fill(sheet, header: list[str], rows: list[list]) -> None:
    sheet.append(header)
    for row in rows:
        sheet.append(row)


def build_export(session: Session) -> bytes:
    """A fresh workbook reflecting the live ledger: README, Inventory, Forecasts, Gaps."""
    site_names = {site.id: site.name for site in session.exec(select(Site)).all()}
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    workbook = Workbook()

    readme = workbook.active
    readme.title = "README"
    readme["A1"] = "ReliefLink live export"
    readme["A2"] = f"Generated {stamp}"
    readme["A3"] = EXPORT_NOTE
    readme.column_dimensions["A"].width = 110

    _fill(
        workbook.create_sheet("Inventory"),
        ["site", "category", "count", "confidence", "source", "updated"],
        [
            [
                site_names.get(s.site_id, s.site_id),
                s.category,
                s.count,
                s.confidence,
                s.source,
                str(s.created_at),
            ]
            for s in sorted(latest_snapshots(session), key=lambda s: (s.site_id, s.category))
        ],
    )
    _fill(
        workbook.create_sheet("Forecasts"),
        ["site", "category", "predicted_demand", "multiplier", "horizon_hours", "reason"],
        [
            [
                site_names.get(f.site_id, f.site_id),
                f.category,
                f.predicted_demand,
                f.multiplier,
                f.horizon_hours,
                f.reason,
            ]
            for f in sorted(latest_forecasts(session), key=lambda f: (f.site_id, f.category))
        ],
    )
    _fill(
        workbook.create_sheet("Gaps"),
        ["site", "category", "current", "predicted_demand", "gap (positive = shortage)"],
        [
            [
                site_names.get(g["site_id"], g["site_id"]),
                g["category"],
                g["current"],
                g["predicted_demand"],
                g["gap"],
            ]
            for g in sorted(compute_gaps(session), key=lambda g: -g["gap"])
        ],
    )
    _fill(
        workbook.create_sheet("Upload"),
        ["site", "category", "count", "county", "lat", "lon"],
        [["Your Pantry Name", category, 0, "", "", ""] for category in CATEGORIES],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_template() -> bytes:
    """A starter sheet a partner can fill in and upload."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Upload"
    _fill(
        sheet,
        ["site", "category", "count", "county", "lat", "lon"],
        [
            ["Example Pantry", "canned_goods", 120, "Alameda", 37.80, -122.27],
            ["Example Pantry", "produce", 40, "Alameda", 37.80, -122.27],
            ["Example Pantry", "dairy", 25, "Alameda", 37.80, -122.27],
            ["Example Pantry", "dry_goods", 90, "Alameda", 37.80, -122.27],
        ],
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
