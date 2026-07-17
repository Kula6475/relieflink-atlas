"""Tests for the spreadsheet import/export bridge. Run with: pytest"""

import io
import os
import tempfile

os.environ.setdefault("LEDGER_DB", os.path.join(tempfile.mkdtemp(), "test.db"))

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from ledger.main import app  # noqa: E402

client = TestClient(app)

XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def xlsx_bytes(rows: list[list]) -> bytes:
    workbook = Workbook()
    for row in rows:
        workbook.active.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_import_creates_site_and_snapshots():
    payload = xlsx_bytes(
        [
            ["Site Name", "Category", "Qty", "County"],
            ["Fremont Family Pantry", "Canned Goods", 77, "Alameda"],
            ["Fremont Family Pantry", "produce", 33, "Alameda"],
        ]
    )
    response = client.post(
        "/spreadsheets/import",
        files={"file": ("inventory.xlsx", payload, XLSX_TYPE)},
    )
    assert response.status_code == 200
    summary = response.json()
    assert summary["imported_rows"] == 2
    assert "Fremont Family Pantry" in summary["created_sites"]
    assert summary["export_url"] == "/spreadsheets/export"

    sites = client.get("/sites").json()
    site = next(s for s in sites if s["name"] == "Fremont Family Pantry")
    inventory = client.get("/inventory", params={"site_id": site["id"]}).json()
    counts = {row["category"]: row["count"] for row in inventory}
    assert counts == {"canned_goods": 77, "produce": 33}
    assert all(row["source"] == "spreadsheet" for row in inventory)


def test_import_csv_and_bad_rows_are_reported():
    csv_data = (
        "site,category,count\n"
        "CSV Pantry,dry goods,12\n"
        "CSV Pantry,mystery meat,5\n"
        "CSV Pantry,dairy,not_a_number\n"
    ).encode()
    response = client.post(
        "/spreadsheets/import", files={"file": ("inventory.csv", csv_data, "text/csv")}
    )
    assert response.status_code == 200
    summary = response.json()
    assert summary["imported_rows"] == 1
    assert len(summary["skipped"]) == 2


def test_import_rejects_empty_upload():
    response = client.post("/spreadsheets/import", files={"file": ("empty.csv", b"", "text/csv")})
    assert response.status_code == 422


def test_export_is_a_live_workbook():
    response = client.get("/spreadsheets/export")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_TYPE)

    workbook = load_workbook(io.BytesIO(response.content))
    assert {"README", "Inventory", "Forecasts", "Gaps", "Upload"} <= set(workbook.sheetnames)


def test_template_downloads():
    response = client.get("/spreadsheets/template")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    header = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]
    assert header[:3] == ["site", "category", "count"]
